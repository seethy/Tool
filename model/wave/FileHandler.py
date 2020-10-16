import openpyxl
import os
import model.math.NewtonMethod as nm

class FileHandler():
    def BatchTxtCal(self, srcDir, destdir):
        if os.path.isdir(srcDir):
            for file in os.listdir(srcDir):
                filepath = srcDir + os.sep + file
                if os.path.isfile(filepath):
                    fp = open(filepath, "r")
                    fpw = open(destdir + os.sep + file + ".o", "w")
                    while True:
                        line = fp.readline()
                        line = line.replace("\t", " ")
                        if not line:
                            break
                        else:
                            deep = float(line.split(" ")[0].strip())
                            period = float(line.split(" ")[1].strip())
                            newtonmethod = nm.NewtonMethod()
                            if deep > 0 and period > 0:
                                wavelength = newtonmethod.GetWaveLenth(period=period, depth = deep)
                                fpw.write('%-10.2f%-10.2f%-10.2f\n' % (deep, period, wavelength))
                            else:
                                fpw.write('%-10.2f%-10.2f%s\n' % (deep, period, ''))
                    fp.close()
                    fpw.flush()
                    fpw.close()

    def BatchExcelCal(self, fileName, deepCol, periodCol, lengthCol):
        workbook = openpyxl.load_workbook(fileName)

        for sheetname in workbook.get_sheet_names():
            sheet = workbook[sheetname]
            for r in range(1, len(sheet.rows) + 1):
                deepvalue = float(sheet.cell('%s%s' % (deepCol, r)).value)
                periodvalue = float(sheet.cell('%s%s' % (periodCol, r)).value)
                newtonmethod = nm.NewtonMethod()
                if deepvalue> 0 and periodvalue > 0:
                    wavelength = newtonmethod.GetWaveLenth(period=periodvalue, depth=deepvalue)
                    sheet.cell('%s%s' % (lengthCol, r)).value = round(wavelength, 2)
        workbook.save(fileName)

if __name__ == '__main__':
    filehandler = FileHandler()
    filehandler.BatchTxtCal("E:\\test1","E:\\test2")
